import openpyxl


result = {}

# Postive boundary values
result['min'] = False
result['min+'] = False
result['nom'] = False
result['max-'] = False
result['max'] = False

# Negative boundary values
result['min--'] = False
result['min-'] = False
result['max+'] = False
result['max++'] = False

# Other negative values
result['invalid_number_type'] = False
result['invalid_type'] = False
result['null'] = False

filename = 'inputs/birthday_inputs.xlsx'
        
wb = openpyxl.load_workbook(filename)

sheet = wb[wb.sheetnames[0]]
for row in sheet.iter_rows(2):
    
    month = row[0].value
    
    if month is None:
        result['null'] = True
    elif type(month) is float:
        result['invalid_number_type'] = True
    elif type(month) is not int:
        result['invalid_type'] = True
    elif month == 0:
        result['min'] = True
    elif month == 1:
        result['min+'] = True
    elif month > 1 and month < 11:
        result['nom'] = True
    elif month == 11:
        result['max-'] = True
    elif month == 12:
        result['max'] = True
    elif month < -1:
        result['min--'] = True
    elif month == -1:
        result['min-'] = True
    elif month == 13:
        result['max+'] = True
    elif month > 13:
        result['max++'] = True
        
        
print('\n\n----------------------------------')
print('------ COVERED CATEGORIES -------')
print('----------------------------------')
        
for key,val in result.items():
    if val:
        print(f'Great job. Category "{key}" was covered.')
        
        
print('\n\n----------------------------------')
print('------ MISSED CATEGORIES --------')
print('----------------------------------')     
        
for key,val in result.items():
    if not val:
        print(f'Category "{key}" was not covered!')
        
print('\n\n')
            
            
            

            
        
        
        